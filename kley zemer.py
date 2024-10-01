import openpyxl
from selenium import webdriver
from selenium.common import TimeoutException, NoSuchElementException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()

driver.get("https://www.kley-zemer.co.il/")

driver.maximize_window()

wait = WebDriverWait(driver, 10)

try:
    search_box = wait.until(EC.presence_of_element_located((By.ID, "searchbox")))

    search_box.clear()

    search_box.send_keys("bass + guitar")

    search_box.send_keys(Keys.ENTER)

except TimeoutException:
    print("Search box not found. Skipping search.")

try:
    search_results = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='col-xs-6 col-sm-3 col-md-2']")))
except TimeoutException:
    print("Search results not found. Website may be loading content dynamically.")

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1).value = "Description"
sheet.cell(row=1, column=2).value = "Price"

for index, search_result in enumerate(search_results, start=2):
    try:
        description_element = search_result.find_element(By.XPATH, ".//div[@class='description']/h2/div[2]")
        description = description_element.text

        price_element = search_result.find_element(By.XPATH, ".//div[@class='text-center priceBlock']/span[@class='oldprice']")
        price = price_element.text

        sheet.cell(row=index, column=1).value = description
        sheet.cell(row=index, column=2).value = price

    except NoSuchElementException:
        print(f"Search result {index-1} may not have the expected elements.")
        continue

workbook.save("C:\\Users\\shomrat\\PycharmProjects\\pythonProject\\bass_guitar_list.xlsx")

driver.quit()
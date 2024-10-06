import openpyxl

def save_to_excel(data, filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "rooms_list"

    header_row = ["Room Name", "Price", "Amenities"]
    for col_num, header in enumerate(header_row, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, room_info in enumerate(data, start=2):
        for col_num, key in enumerate(header_row, start=1):
            sheet.cell(row=row_num, column=col_num, value=room_info[key])

    workbook.save(filename)
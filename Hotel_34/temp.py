import time
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from datetime import datetime, timedelta
import random

driver = webdriver.Chrome()

driver.get("https://hotelh34.com/")
driver.maximize_window()

order_button = driver.find_element(By.PARTIAL_LINK_TEXT, "הזמינו עכשיו")
order_button.click()
time.sleep(2)
tabs = driver.window_handles
driver.switch_to.window(tabs[1])

accept_cookies = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//div[@id='klaro-cookie-notice']//button[text()='אישור']"))
)
accept_cookies.click()


def wait_for_calendar(driver):
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".sb-checkin.month"))
    )


def is_date_available(driver, date):
    month_select = Select(driver.find_element(By.CSS_SELECTOR, ".sb-checkin.month"))
    month_select.select_by_value(f"{date.month},{date.year}")

    time.sleep(1)

    day_elements = driver.find_elements(By.CSS_SELECTOR, ".rg-day")
    for day_element in day_elements:
        day_header = day_element.find_element(By.CSS_SELECTOR, ".rg-day-header")
        if day_header.text == str(date.day):
            return "no-availability" not in day_element.get_attribute("class")
    return False


def select_random_dates(driver):
    wait_for_calendar(driver)
    today = datetime.now()
    max_date = today + timedelta(days=365)  # Assuming bookings are allowed up to 1 year in advance

    attempts = 0
    max_attempts = 50

    while attempts < max_attempts:
        check_in = today + timedelta(days=random.randint(1, 365))
        check_out = check_in + timedelta(days=7)

        if check_out <= max_date and is_date_available(driver, check_in) and is_date_available(driver, check_out):
            return check_in, check_out

        attempts += 1

    raise Exception("Unable to find available dates after multiple attempts")


def select_date(driver, date, is_checkin):
    selector = ".sb-checkin.month" if is_checkin else ".sb-checkout.month"
    month_select = Select(driver.find_element(By.CSS_SELECTOR, selector))
    month_select.select_by_value(f"{date.month},{date.year}")

    time.sleep(1)

    day_elements = driver.find_elements(By.CSS_SELECTOR, ".rg-day")
    for day_element in day_elements:
        day_header = day_element.find_element(By.CSS_SELECTOR, ".rg-day-header")
        if day_header.text == str(date.day):
            day_element.click()
            break

    if is_checkin:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".sb-checkout.month"))
        )


check_in, check_out = select_random_dates(driver)

select_date(driver, check_in, True)

select_date(driver, check_out, False)

check_availability = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, ".sb-do-search-cmd")))
check_availability.click()

time.sleep(3)
search_results = WebDriverWait(driver, 30).until(
    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".sb-room-container.closed"))
)

for result in search_results:
    room_name_element = result.find_element(By.CSS_SELECTOR, ".sb-room-card__name")
    room_name = room_name_element.text.strip()

    price_element = result.find_element(By.CSS_SELECTOR, ".sb-room-card__price .sale-price-big")
    price = price_element.text.strip()

    print(f"Room Name: {room_name}")
    print(f"Price: {price}")
    print("-" * 50)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "rooms_list"

    header_row = ["Room Name", "Price", "Amenities"]
    for col_num, header in enumerate(header_row, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    row_num = 2
    for result in search_results:
        room_name_element = result.find_element(By.CSS_SELECTOR, ".sb-room-card__name")
        room_name = room_name_element.text.strip()

        price_element = result.find_element(By.CSS_SELECTOR, ".sb-room-card__price .sale-price-big")
        price = price_element.text.strip()

        amenities_elements = result.find_elements(By.CSS_SELECTOR, ".sb-icon-amenities")
        amenities = ", ".join([amenity.get_attribute("title") for amenity in amenities_elements])

        sheet.cell(row=row_num, column=1, value=room_name)
        sheet.cell(row=row_num, column=2, value=price)
        sheet.cell(row=row_num, column=3, value=amenities)
        row_num += 1

    workbook.save("rooms_list.xlsx")
input('x')
